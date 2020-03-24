#################################################################
#                                                               #
#     The user should use the provided template excel file.     #
#                                                               #
#     !! IMPORTANT, the archive sheet (Arkiv) should contain    #
#     all the children currently in the class. This list will   #
#     be used to make the new friend groups                     #
#                                                               #
#################################################################


import math
import sys
import random
from openpyxl import load_workbook

#################################################################
#                                                               #
#                   Checking the infile                         #
#                                                               #
#################################################################


# Check for right number of arguments
if len(sys.argv) != 2:
    print("Usage: python vennegruppe.py groups.xlsx")
    sys.exit(1)

# Open infile
print("Opening infile")
infile = sys.argv[1]
wb_infile = load_workbook(filename=infile, read_only=True)

# Check if the provided set-up excel file is used by checking sheet names
sheet1, sheet2 = wb_infile.sheetnames
if not sheet1 == 'Vennegruppe' or not sheet2 == 'Arkiv':
    print("Please use the provided example file")
    sys.exit(1)

#################################################################
#                                                               #
#           Read and store the previous groups                  #
#                                                               #
#################################################################


# grab the groups worksheet (Vennegruppe)
sheet1_infile = wb_infile["Vennegruppe"]

# Initialize dictionary to hold the previous friend groups (Vennegruppe)
groups = {}
number_of_groups = 0
number_of_kids_in_group = 0

# grab the latest friend groups (Vennegruppe)
print("\nReading latest friend groups")
while True:
    number_of_groups += 1

    if not sheet1_infile.cell(row=1, column=number_of_groups).value:
        break

    groups[number_of_groups] = list()

    # Grab the names of the children in the group
    while True:
        number_of_kids_in_group += 1

        if not sheet1_infile.cell(row=(number_of_kids_in_group + 1), column=number_of_groups).value:
            number_of_kids_in_group = 0
            break

        groups[number_of_groups].append(sheet1_infile.cell(row=(number_of_kids_in_group + 1),
                                                           column=number_of_groups).value)

#################################################################
#                                                               #
#   Store history of 'who_has_been_in_a_group_with' in a dict   #
#                                                               #
#################################################################


# Change to archive sheet (Arkiv)
sheet2_infile = wb_infile["Arkiv"]

# Initialize variables to store 'who_has_been_in_a_group_with'
archive = {}
child = 0
child_name = ""
has_been_in_group_with = list()
iterator = 3  # First column of the many "has_been_in_group_with" columns

# grab the name and gender of all the children (Kjønn, Navn)
print("Storing all the previous friend groups")
while True:
    child += 1

    # breaks out of the while loop when the next cell is empty (There are no more children)
    if not sheet2_infile.cell(row=(child + 5), column=1).value:
        break

    # grabs the child's name and stores it in a variable
    child_name = sheet2_infile.cell(row=(child + 5), column=1).value

    # grab all the names of "has_been_in_group_with" (Har vært i gruppen med)
    while True:

        # breaks out of the while loop when the next cell is empty (There are no more names in "has_been_in_group_with")
        if not sheet2_infile.cell(row=(child + 5), column=iterator).value:
            iterator = 3
            break

        # Grabs and stores the child names and stores in a list
        has_been_in_group_with.append(sheet2_infile.cell(row=(child + 5), column=iterator).value)
        iterator += 1

    # After the list "has_been_in_group_with" is done, this list is added to the archive the main child
    archive[child_name] = {"gender": sheet2_infile.cell(row=(child + 5), column=2).value,
                           "has_been_in_group_with": has_been_in_group_with}

    # List is emptied for the next main child
    has_been_in_group_with = list()

# Close the excel file
wb_infile.close()

#################################################################
#                                                               #
#       Add latest groups to the "has_been_in_group_with"       #
#                                                               #
#################################################################


# Access each group from the latest groups
print("Adding the latest groups to the archive")
for group in groups:

    # Access the names in the group
    for name in groups[group]:

        # Check if name is in the archive
        if name in archive.keys():

            # Add the other children of the latest group into the archive
            for kid in groups[group]:
                if name != kid:
                    archive[name]["has_been_in_group_with"].append(kid)

#################################################################
#                                                               #
#                   Make new random groups                      #
#                                                               #
#################################################################


# Make a list containing all the names
list_of_names = list(archive.keys())
randomized_list_of_names = list_of_names.copy()


# Function to randomize the order of the names in the list
def randomize_children():
    random.shuffle(randomized_list_of_names)
    return randomized_list_of_names


# Initializing variables to make new friend groups
total_nr_of_groups = 1
nr_of_5_kids_per_group = len(archive) % 4
nr_of_4_kids_per_group = int((len(archive) - nr_of_5_kids_per_group * 5) / 4)
new_friend_groups = {}
genders_per_group = {}

# Calculate how many groups are needed
length_archive = len(archive)
if length_archive >= 15:
    total_nr_of_groups = math.floor(len(archive) / 4)
else:
    if length_archive % 4 <= 2:
        total_nr_of_groups = math.floor(len(archive) / 4)
    else:
        total_nr_of_groups = math.ceil(len(archive) / 4)


# Print a summary of the total number of children and the group sizes that will be made
print("\nTotal number of children: {0}\nTotal number of groups: {1} \n {2} groups with 5 children and,"
      "\n {3} groups with 4 children".format(length_archive, nr_of_5_kids_per_group + nr_of_4_kids_per_group,
                                             nr_of_5_kids_per_group, nr_of_4_kids_per_group))


# Function to make new groups from the randomized list
def make_groups():
    new_friend_groups.clear()
    random_list_of_names = randomize_children()
    i = 1

    # Make the groups with 5 children
    for group_with_five in range(0, (nr_of_5_kids_per_group * 5), 5):
        # Assigns the friend group number and store the names
        new_friend_groups[i] = random_list_of_names[group_with_five:(group_with_five + 5)]

        # iterator + 1 for the next group (if any)
        i += 1

    # Make the groups with 4 children
    for group_with_four in range((nr_of_5_kids_per_group * 5),
                                 (nr_of_4_kids_per_group * 4 + nr_of_5_kids_per_group * 5), 4):
        # Assigns the friend group number and store the names
        new_friend_groups[i] = random_list_of_names[group_with_four:(group_with_four + 4)]

        # iterator + 1 for the next group (if any)    
        i += 1

    return new_friend_groups


# Function to make a list of all the genders in one group
def list_of_genders():
    genders_per_group.clear()
    new_friend_groups_genders = make_groups()
    tmp_genders = list()  # empty list to hold the genders of the children in one group
    j = 0

    for group_genders in new_friend_groups_genders:
        for name_genders in new_friend_groups_genders[group_genders]:
            tmp_genders.append(archive[name_genders]["gender"])

        genders_per_group[j] = tmp_genders
        tmp_genders = list()
        j += 1

    return genders_per_group, new_friend_groups_genders


# Function to score the new friend group
#   10 points for only one girl or boy in a group
#   50 point when group consists of only girls or boys
#   100 points for having been in same group before
def scoring():
    score = 0
    genders_per_group_scoring, new_friend_groups_genders_scoring = list_of_genders()

    # Iterate through the genders of the newly made groups
    for group_scoring in genders_per_group_scoring:
        girl = 0
        boy = 0

        # Count the number of girls (jente) and boys (gutt) in each group
        for gender_scoring in genders_per_group_scoring[group_scoring]:
            if gender_scoring == "jente":
                girl += 1
            else:
                boy += 1

        # Score the group if they have only one or no girls/boys
        if girl or boy <= 1:
            if girl or boy == 1:
                score += 10
            else:
                score += 50

    # Iterate through the names of the newly made groups
    for group_score in new_friend_groups_genders_scoring:
        for name_score in new_friend_groups_genders_scoring[group_score]:
            x = 0
            for archive_name in archive[name_score]["has_been_in_group_with"]:
                while len(new_friend_groups_genders_scoring[group_score]) > x:
                    if archive_name == new_friend_groups_genders_scoring[group_score][x]:
                        score += 100
                    x += 1

    return score, new_friend_groups_genders_scoring


# Initializing variable to score and store each newly made friend group
friend_group_options = {}


def group_options(how_many_options):
    y = how_many_options
    friend_group_options.clear()

    for option in range(y):
        group_score, new_friend_groups_option = scoring()
        tmp = "group " + str(option + 1)
        friend_group_options[tmp] = {}
        for key, value in new_friend_groups_option.items():
            friend_group_options[tmp][key] = value
        friend_group_options[tmp]['score'] = group_score

    return friend_group_options


def best_scoring_random_made_group():
    # print("\nThe more random groups this program is allowed to create, "
    #       "the more likely is it that a group with at least one girl or boy"
    #       "and is the least possible overlap in children who have been in a group together before"
    #       "is created."
    #       "\nHowever, too many groups and you'll have to wait a long time for your result"
    #       "Please enter a number of groups the program can create to get the best result")

    user_input = 10000  # Can be changed to user input with commend above and : int(input("\nHow many random groups: "))

    # Inform the user that the new groups are being made
    print("\nCreating new friend groups")

    friend_group_options_final = group_options(user_input)

    # Find the lowest scoring group
    lowest_scoring_group_number = ''
    lowest_score = float('inf')
    for keys in friend_group_options_final.keys():
        if friend_group_options_final[keys]["score"] < lowest_score:
            lowest_score = friend_group_options_final[keys]["score"]
            lowest_scoring_group_number = keys

    # Print the new friend group to the screen
    for k, v in friend_group_options_final[lowest_scoring_group_number].items():
        if k == "score":
            print("\nWith a total score of:", v)  # Will be deleted at some point and do nothing instead
        else:
            print("\nVennegruppe", k, end=":  ")
            for kid_name in v:
                print(kid_name, end=", ")

    return friend_group_options_final[lowest_scoring_group_number]


# Function to store the groups in a excel file
def write_groups_in_excel(f_group):
    final_group_to_excel = f_group

    # Open outfile
    print("\nOpening outfile")
    outfile = "new_groups.xlsx"
    wb_outfile = load_workbook(filename=outfile)

    print("file opened")

    # Check if the provided set-up excel file is used by checking sheet names
    sheet1_out, sheet2_out = wb_outfile.sheetnames
    if not sheet1_out == 'Vennegruppe' or not sheet2_out == 'Arkiv':
        print("Error: The outfile could not be opened")
        sys.exit(1)

    # grab the groups worksheet (Vennegruppe)
    sheet1_outfile = wb_outfile["Vennegruppe"]

    # write the latest friend groups (Vennegruppe)
    print("Storing the new friend groups")

    for key_number, final_names in final_group_to_excel.items():

        # The score key-value is not used, and will be ignored
        if not key_number == "score":

            # Print the vennegruppe numbers in excel
            sheet1_outfile.cell(row=1, column=key_number).value = "Vennegruppe " + str(key_number)

            # Iterate through the rows in excel and the names in each group for printing in excel
            row = 2
            for final_name in final_names:
                sheet1_outfile.cell(row=row, column=key_number).value = final_name
                row += 1

                # Delete the fifth name is this groups contains only four children
                # (in case there is a name there from before)
                if len(final_names) <= 4:
                    sheet1_outfile.cell(row=6, column=key_number).value = None

    # Change to archive sheet (Arkiv)
    sheet2_outfile = wb_outfile["Arkiv"]
    print("Updating the archive")

    row_archive = 6

    for archive_name, archive_list in archive.items():
        column_archive = 3

        # Write the name of the child in the first column
        sheet2_outfile.cell(row=row_archive, column=1).value = archive_name
        row_archive += 1

        # Read the genders, and list of has_been_in_group_with
        for archive_key, archive_value in archive_list.items():

            # Write gender in the excel file
            if archive_key == "gender":
                sheet2_outfile.cell(row=row_archive-1, column=2).value = archive_value

            # Write all the names from has_been_in_group_with in the excel file
            else:
                for archive_has_been_in_group_with in archive_value:
                    sheet2_outfile.cell(row=row_archive-1, column=column_archive).value = archive_has_been_in_group_with
                    column_archive += 1

    # Safe the outfile
    wb_outfile.save("groups.xlsx")

    print("\nFile saved")


# Function to ask the user if the created friend groups are accepted
def accept():
    accept_group = ''
    final_group = best_scoring_random_made_group()

    while True:
        if accept_group.lower() in ('y', 'yes', 'n', 'no'):
            break
        else:
            accept_group = input("\nDo you accept these vennegruppe? y/n ")

    if accept_group.lower() in ('n', 'no'):
        accept()
    elif accept_group.lower() in ('y', 'yes'):
        write_groups_in_excel(final_group)
    else:
        print("Something went wrong with saving the groups, please try again")
        sys.exit(1)


accept()

# Success
sys.exit(0)

#################################################################
#                                                               #
#               Written by Hanna Noordzij 2019                  #
#                                                               #
#################################################################
